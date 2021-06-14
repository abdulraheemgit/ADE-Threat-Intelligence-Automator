# -*- coding: cp1252 -*-
from __future__ import print_function
import simplejson
import urllib
import urllib2
import ssl
import os
import ipaddress
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Border, Side, Color, colors

def copyRights():
    print (", ,-.    ,-.              .            	")
    print ("| |  )   |  )             |            	")
    print ("| |-'    |-<  ,-. ,-. ,-. | . , ,-. ;-.	")
    print ("| |      |  \ |-' `-. | | | |/  |-' |  	")
    print ("' '      '   '`-' `-' `-' ' '   `-' '  	")
    print ("Developed by AR (Abdul Raheem) 				 ")
    print ("Powerd by SL CERT|CC			             ")
    print ("Copyright © 2018 AR. All Rights Reserved.    ")
    
def deleteFile(confData):
    try:
        os.remove(confData["filePath"])
        return True
    except:
        return False

def readConf(fileType=None):                                                            #reads the configuration details in ADE.config                                                    
    confData = {}
    if fileType == "conficker":
        read = 'Config/conficker.config'
    if fileType == "nitol":
        read = 'Config/nitol.config'
    else:
        read = 'Config/ADE.config'        
    readFile = open (read, "r")
    data = readFile.read()
    readFile.close()
    for line in data.split('\n'):
        if line:
            if line[0] == "@":
                col = line[1:].split('::')
                if not col[1]:
                    col.append('')                    
                confData[col[0].strip()] = col[1].strip()
    return confData

def download(confData):
    flag = 0
    #setting parameter for file download
    parameters = {"apikey": confData["apiKey"],
                  "fileid": confData["fileName"]
                  }
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE                                                 #disable SSL (not to veify certification - self signed)

    dateStripped = str(confData["fileDate"]).split()[0]
    downloadedFile = r''+confData["fileName"]+'-'+dateStripped+'('+confData["dataType"]+').csv.asc'                         #renaming the file with issue date
    print ("downloading file - "+downloadedFile+"... ", end='')
    
    data = urllib.urlencode(parameters)
    req = urllib2.Request(confData["downloadUrl"], data)                            #sending the request
    
    response = urllib2.urlopen(req, context=ctx)
    json = response.read()                                                          #reading the response
    

    if not os.path.exists(confData['downloadFolder']+str(datetime.now().date())):   #creates the download folder
        os.makedirs(confData['downloadFolder']+str(datetime.now().date()))
    folder = confData['downloadFolder']+ str(datetime.now().date())
                
    try:
        json_dict = simplejson.loads(json)
        json = simplejson.dumps(json_dict, indent=4)
    except:
        pass
    finally:        
        f2 = open(confData["downloadLogsFolder"]+confData["downloadLogsFile"], "a+")                                        #open to log file
        if json:
            flag = 1
            f = open(folder+"/"+downloadedFile, "wb")                                                                       #open file to write downloaded content
            f.write(json)       
            f.close()            
            f2.write(downloadedFile + " on "+str(datetime.now())+"\n")                                                      #log if download succeeded
            print ("done!")
        else:
            f2.write(downloadedFile + " download failed - "+str(datetime.now())+"\n")                                       #log if download failed
            print ("failed!")
        f2.close()
    return [flag, folder+"/"+downloadedFile]

def escapeString(string, confData):                                                 #escapes special charactors and sanitize string
    escape = None   
    if confData['escape']:
        escape = confData['escape'].strip()            
    return string.strip(escape)

def ispFound(ip, ispIp):
    result = False
    isp = ''
    keyIp = ''
    for netIp, values in ispIp.iteritems():                                         #iterate the list of ip addresses of ISP's        
        uIp = unicode(netIp+"/"+values[0])                                          #convert to unicode
        ip = ipaddress.ip_address(unicode(ip))
        try:
            result = ip in ipaddress.ip_network(uIp)                                #check if ip within the net range of ISP
            if result:
                keyIp = netIp
                ispIp[netIp][3] += 1                                                #increse ip incident count/isp has multiple ip
                break
        except ValueError:
            print ("inappropriate network address")                                 #exception handling for incorrect ISP's network address format
            print ("please review the network address of " + values[1])
            print ("makesure no host bits are set in the network ip")            
            isp = values[1]    
    return [result, isp, ispIp, keyIp]
def checkRedundancy(ip, confData):                                                  #check if th eip is in the previous reported list
    f = open(confData["redundentIp"].strip(), "r")                                  #opens the ip log file
    data = f.read()
    f.close()
    if ip.strip() in data: return True                                              #check if ip in log file
    else: return False

def writeToLog(ip,confData):
    f = open(confData["redundentIp"].strip(), "a+")
    f.write(ip+"\n")
    f.close()
    
def createFile(confData, files):
    isp = load_workbook(confData['ispFile'])                                        #load excel file with ISP details
    wsIp = isp[confData['ispIp']]                                                   #select sheet for IP details
    wsContact = isp[confData['ispContact']]                                         #select sheet for contact details
    logType = ''
    isp = {}
    ispIp = {} #ispIp
    unknownIp = []
    u=0    
    for i in  range(2,wsIp.max_row+1):                                              #reading isp ip info from ispContact.xlsx
        if wsIp[i]:
            #{ipblock : [subnet,isp,found/notfound, datacount]}
            ispIp[str(wsIp.cell(row=i,column=3).value)] = [str(wsIp.cell(row=i,column=4).value),str(wsIp.cell(row=i,column=5).value), 0,0] 
           
    for i in  range(2,wsContact.max_row+1):                                         #reading isp contact info from ispContact.xlsx
        if wsIp[i]:
            #{isp : [totalcount,contact,email,file path,[logtype] ]}
            isp[str(wsContact.cell(row=i,column=2).value)] = [0,str(wsContact.cell(row=i,column=3).value),str(wsContact.cell(row=i,column=4).value),'',[]]
    createdFile = []                                                                # = [ [isp,logType] ]
    for filePath in files:
        if 'conficker' in filePath:
            logType = "conficker"
            try:
                confDataF = readConf("conficker")
            except:
                print ("Error in conficker.config file")
                raw_input("Hit enter to exit the program")
                exit()
        elif 'nitol' in filePath:
            logType = "nitol"
            try:
                confDataF = readConf("nitol")
            except:
                print ("Error in nitol.config file")
                raw_input("Hit enter to exit the program")
                exit()
        else:
            continue
        readFile = open (filePath, "r")
        data = readFile.read()                                                      #read the input file
        readFile.close()
        data = escapeString(data, confData)
        for line in data.split('\n'):
            if line:
                if 'ip' in line.lower() and 'urlpath' in line.lower() and 'country' in line.lower():                        #skip heading
                    continue
                if 'client_ip' in line.lower() and 'server_port' in line.lower() and 'country' in line.lower():             #skip heading
                    continue
                ipCol = int(confDataF['ipCol'])
              
                cols = line.split(",")
                if checkRedundancy(cols[ipCol], confData):                          #check if ip is reported recently
                    continue

                numRange = '([01]?[0-9]?[0-9]|2[0-4][0-9]|25[0-5])'
                numRangeStart = '([1-9]|[1-9][0-9]|1[0-9][0-9]|2[0-4][0-9]|25[0-5])'
                regEx = re.compile('^'+numRangeStart+'((\.'+numRange+'){0,3})$')    #regex to check ip format

                NoneType = type(None)
                if type(regEx.match(cols[ipCol])) is NoneType:                      #check ip format
                    print ("Skipping invalid IP - " + cols[ipCol])
                    continue

    
                if checkRedundancy(cols[ipCol], confData):                          #check if ip is reported recently
                    continue
                ipFound = ispFound(cols[ipCol].strip(), ispIp)
                
                ispIp = ipFound[2]
                if ipFound[0]:                                                      #check if incident ip is within the isp ip range list
                    writeToLog(cols[ipCol],confData)                    
                    wbIsp = ispIp[ipFound[3].strip()][1]                            #if yes, get the isp name of the above ip
                    if not isp[wbIsp][0] > 0:                        
                        exec('globals()["'+wbIsp+'_'+logType+'"] = Workbook()')     #creating excel
                        exec('ws'+ wbIsp + '_' + logType +' = '+ wbIsp + '_' + logType +'.active')
                        if not logType in isp[wbIsp][4]:
                            isp[wbIsp][4].append(logType)                           #enter log type under isp if not listed
                            createdFile.append([wbIsp, logType])                            
                        first = ''
                        last = ''
                        c = 1
                        for head in confDataF['colunmHeadings'].split(','):
                            exec('ws'+ wbIsp + '_' + logType +'.cell(row=1, column='+str(c)+', value="'+head+'")')          #write heading in excel                        
                            exec('row = ws'+ wbIsp + '_' + logType +'["'+chr(64+c)+'1"]');row.font = Font(bold=True)        #applying style to heading
                            c +=1
                        c = 1                    
                        for colNum in confDataF['columnNumbers'].split(','):
                            exec('ws'+ wbIsp + '_' + logType +'.cell(row=2, column='+str(c)+', value="'+cols[int(colNum)]+'")')             #write data to excel
                            c +=1                                                                                 
                        isp[wbIsp][0] += 1                                          #increse total incident count for ISP
                        
                    else:
                        if not [wbIsp, logType] in createdFile:
                            exec('globals()["'+wbIsp+'_'+logType+'"] = Workbook()') #creating excel
                            exec('ws'+ wbIsp + '_' + logType +' = '+ wbIsp + '_' + logType +'.active')
                            createdFile.append([wbIsp, logType])
                            c = 1
                            for head in confDataF['colunmHeadings'].split(','):
                                exec ('ws' + wbIsp + '_' + logType + '.cell(row=1, column=' + str(c) + ', value="' + head + '")')           # write heading in excel
                                exec ('row = ws' + wbIsp + '_' + logType + '["' + chr(64 + c) + '1"]');row.font = Font(bold=True)           # applying style to heading
                                c += 1
                            c = 1
                        isp[wbIsp][0] += 1                                          #increse total incident count for ISP
                        dataCount = ispIp[ipFound[3]][3]                            #to map the excel rows                   
                        c = 1
                        for colNum in confDataF['columnNumbers'].split(','):                        
                            exec('ws'+ wbIsp + '_' + logType +'.cell(row='+str(dataCount+1)+', column='+str(c)+', value="'+cols[int(colNum)]+'")')      #write data to excel
                            c +=1           
                else:
                    unknownIp.insert(u, escapeString(line, confData).split(",")[0]) #logs if there is any unknown ip                
                    u +=1
    ipCheck = False
    for isp1,count in isp.iteritems():        
        if count[0] > 0:
            ipCheck = True                   
            time = str(datetime.now().time())[0:2]+"HRS "+str(datetime.now().time())[3:5]+"MIN"
            if not os.path.exists(confData['outputFolder']+str(datetime.now().date())+'  '+time+'/'+isp1):                  #creates the output folder
                os.makedirs(confData['outputFolder']+str(datetime.now().date())+'  '+time+'/'+isp1)
            folder = confData['outputFolder']+ str(datetime.now().date())+'  '+time+'/'+isp1+'/'                            #creates the output folder
            isp[isp1][3] =  folder
            for ispFile in createdFile:
                if ispFile[0] == isp1:
                    path = folder
                    path += isp1 + '_' + ispFile[1]
                    print ('creating file '+isp1+ '_' + ispFile[1]+'.xlsx...')
                    exec(isp1 + '_' + ispFile[1] +'.save("'+path+'.xlsx")')         #saves the excel file
                    exec(isp1+ '_' + ispFile[1] +'.close()')

    if not ipCheck:
        print ("no new data to communicate.\nall ip addresses are reported within the last "+confData["redundencyPeriod"]+" week(s)\n")
        raw_input("press enter to exit the program")
        exit()
    else:
        print ('\nfiles are created! \nplease review it before sending the email!!!\n')
    return [isp, ispIp, unknownIp]

def generateReport(isp, asn, unknown, sentMailStatus, confData):                    #this function will create a summary report of compromized ip's and its relevent communication
    
    
    report = Workbook()
    ws = report.active
    report.remove_sheet(ws)
    wsReport = report.create_sheet("Compromized IP")                                #create excel sheet  
    wsReport.cell(row=1,column=1,value='#Resolver 1.0.0')
    wsReport.cell(row=2,column=1,value='#Auther : Abdul Raheem')
    wsReport.cell(row=3,column=1,value='#Powered by : SL CERT|CC')
    wsReport.cell(row=5,column=1,value='Report for compromized IP').font = Font(bold=True, size=14)
    
    thinBorder = Border(left=Side(border_style='thin', color='FF000000'),right=Side(border_style='thin',color='FF000000'),  #styles
                top=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'))
    thickBorder = Border(left=Side(border_style='medium', color='FF000000'),right=Side(border_style='medium',color='FF000000'),
                top=Side(border_style='medium',color='FF000000'), bottom=Side(border_style='medium',color='FF000000'))

    totalIP = 0
    row = 9
    row2 = 9  
    if unknown:
        wsReport.cell(row=row2,column=7,value='Unknown IP List').font = Font(bold=True)                                     #logs the unknown IP
        row2 += 1
        for asn1 in unknown:
            wsReport.cell(row=row2,column=7,value=asn1)
            row2 += 1
        wsReport.cell(row=row2,column=7,value='Please find the respective ISP for the above IP(s)')
        row2 += 1
        wsReport.cell(row=row2,column=7,value='You may add the respective ISP details to '+confData["fileName"]+' and run the program again')       
            
        
    wsReport.cell(row=row,column=1,value='ISP').font = Font(bold=True)              #logs the compromized ip details
    wsReport.cell(row=row,column=2,value='IP Range').font = Font(bold=True)
    wsReport.cell(row=row,column=3,value='No of Compromized IPs ').font = Font(bold=True)
    for key,item in isp.iteritems():
        if item[0] > 0:
            row += 1
            for asnKey, data in asn.iteritems():
                if data[1] == key:
                    wsReport.cell(row=row,column=1,value=key).border = thinBorder
                    wsReport.cell(row=row,column=2,value=asnKey+"/"+data[0]).border = thinBorder
                    wsReport.cell(row=row,column=3,value=data[3]).border = thinBorder
                    row += 1
            wsReport.cell(row=row,column=2,value='Total').border = thickBorder        
            wsReport.cell(row=row,column=3,value=item[0]).border = thickBorder
            totalIP +=item[0]
            if item[0] > 0:
                wsReport.cell(row=row,column=3).font = Font(color=colors.RED)
            row += 1
    wsReport.cell(row=7,column=1,value="Total IPs reported").font = Font(color=colors.RED)
    wsReport.cell(row=7,column=3,value=totalIP).font = Font(color=colors.RED)

    row = 9    
    wsReport2 = report.create_sheet("Communication")                                #create excel sheet  
    wsReport2.cell(row=1,column=1,value='#Resolver 1.0.0')
    wsReport2.cell(row=2,column=1,value='#Auther : Abdul Raheem')
    wsReport2.cell(row=3,column=1,value='#Powered by : SL CERT|CC')
    wsReport2.cell(row=5,column=1,value='Data Communicated with ISP').font = Font(bold=True)
                                                                                    #logs the logs the reported ISP details
    wsReport2.cell(row=row,column=1,value='ISP').font = Font(bold=True);wsReport2.cell(row=row,column=1).border = thinBorder
    wsReport2.cell(row=row,column=2,value='Contact Person').font = Font(bold=True);wsReport2.cell(row=row,column=2).border = thinBorder
    wsReport2.cell(row=row,column=3,value='Contact Email').font = Font(bold=True);wsReport2.cell(row=row,column=3).border = thinBorder
    wsReport2.cell(row=row,column=4,value='File Sent').font = Font(bold=True);wsReport2.cell(row=row,column=4).border = thinBorder
    wsReport2.cell(row=row,column=5,value='Status').font = Font(bold=True);wsReport2.cell(row=row,column=5).border = thinBorder
    wsReport2.cell(row=row,column=6,value='Feedback').font = Font(bold=True);wsReport2.cell(row=row,column=6).border = thinBorder
    row += 1

    for ispKey,data1 in isp.iteritems():
        if data1[0] > 0:
            wsReport2.cell(row=row,column=1,value=ispKey).border = thinBorder
            wsReport2.cell(row=row,column=2,value=data1[1]).border = thinBorder
            wsReport2.cell(row=row,column=3,value=data1[2]).border = thinBorder
            wsReport2.cell(row=row,column=4,value=data1[3]+'*').border = thinBorder
            wsReport2.cell(row=row,column=5,value=sentMailStatus[ispKey][0]).border = thinBorder
            wsReport2.cell(row=row,column=6,value=sentMailStatus[ispKey][1]).border = thinBorder
            row += 1
    print ('creating file ADEReport.xlsx...')
    
    if not os.path.exists(confData['reportFolder']+str(datetime.now().date())):     #creates the reports folder
        os.makedirs(confData['reportFolder']+str(datetime.now().date()))
    folder = confData['reportFolder']+ str(datetime.now().date())+'/'
    
    reoprtFile = confData['reportFile']
    if reoprtFile[len(reoprtFile)-5:len(reoprtFile)] == ".xlsx":                    #set the file name and extention for report
        path = folder+confData['reportFile'][0:len(reoprtFile)-5]+".xlsx"
    else:
        path = folder+confData['reportFile']+".xlsx"
    if os.path.exists(path):
        while True:
            changeName = raw_input("file name "+os.path.basename(path)+" already exist. Do you want to change?(Y/N)")
            if changeName.upper() == 'Y':
                newFile = raw_input("Enter name ")
                if newFile[len(newFile)-5:len(newFile)] == ".xlsx":                 #report ovewrite confirmation
                    path = folder+newFile
                else:
                    path = folder+newFile+".xlsx"
                break
            elif changeName.upper() == 'N':
                break

                
    report.save(path)
    report.close()


